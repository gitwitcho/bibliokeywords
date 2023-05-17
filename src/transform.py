import pandas as pd
import numpy as np
import re
import logging
import webcolors

from typing import List, Optional, Union, Dict, Tuple, Any
from IPython.core.display import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

from config import *
from utilities import *

def modify_cols_biblio_df(biblio_df_: pd.DataFrame,
                                     reshape_base: Optional[Reshape] = None,
                                     reshape_filter: Optional[Union[Dict,List]] = None,
                                     keep_bib_src: bool = True
                                     ) -> pd.DataFrame:
    """
    Rename and retain columns in the bibliographic dataset `biblio_df` based on `reshape_base` with its
    preset criteria or on `reshape_filter`, where you define what column to retain and how to rename them.
    
    The reshape_base parameter allows you to specify preset criteria for renaming and retaining the columns 
    in the bibliographic `DataFrame`. The available values for `reshape_base` are defined in the 
    `config.py` file. For each bibliographic database that can be processed (e.g.Scopus, Lens, 
    Dimensions), there are three levels of detail that can be retained: *All*, *Full*, and *Compact*.

    - The "All" level retains all the columns of the raw dataset from the respective bibliographic database:
      `Reshape.SCOPUS_ALL`, `Reshape.LENS_ALL`, `Reshape.DIMS_ALL`.
    - The "Full" level removes certain columns from `biblio_df` that might not be needed for the analysis:
      `Reshape.SCOPUS_FULL`, `Reshape.LENS_FULL`, `Reshape.DIMS_FULL`.
    - The "Compact" level is similar to "Full" and further reduces the number of columns in `biblio_df`:
      `Reshape.SCOPUS_COMPACT`, `Reshape.LENS_COMPACT`, `Reshape.DIMS_COMPACT`.
    
    Alternatively, you can use the `reshape_filter` parameter to specify a list of column names in `biblio_df` 
    that you want to retain. This allows you to manually select the specific columns you need for further analysis.

    By using either `reshape_base` or `reshape_filter`, you can customize the column selection and reshaping process 
    of the bibliographic `DataFrame` based on your requirements.

    Args:
        biblio_df: 
            The input bibliographic `DataFrame`.
        reshape_base:
            If provided, only the columns specified in the `reshape_base` criteria will be retained and renamed 
            according to the predefined reshape structures (e.g. `reshape_struc_dims_compact`) defined in `config.py`.
        reshape_filter:
            A list or dictionary of column names for retaining specific columns. If a `reshape_base`is provided, 
            the `reshape_filter` only retains columns that are present once `reshape_base` has been applied.
            - If `reshape_filter` is a list, only the columns present in the list will be retained.
            - If `reshape_filter` is a dictionary, the keys represent the columns to be retained,
              and the values represent the new column names.
        keep_bib_src:
            Flag indicating whether to retain the 'bib_src' column in the output `DataFrame`. Otherwise
            it will be removed when applying `rehshape_base`and `reshape_column`(unless you provide it in the list).
            The values in `bib_src` are used in some of the other functions. 
    Returns:
        The modified bibliographic `DataFrame`.

    Raises:
        ValueError:
            - If any of the columns specified in `reshape_base` are not found in the `biblio_df`.
            - If any of the columns specified in `reshape_filter` are not found in the `biblio_df`.
    """

    biblio_df = biblio_df_.copy()

    if reshape_base:
        if not all(col in biblio_df.columns for col in reshape_strucs[reshape_base.value].keys()):
            raise ValueError(f"One or more columns not found in biblio_df: {reshape_strucs[reshape_base.value]}")
        
        reshape_columns = list(reshape_strucs[reshape_base.value].keys())

        if ('bib_src' in biblio_df.columns) and ('bib_src' not in reshape_columns):
            reshape_columns.append('bib_src')

        biblio_df = biblio_df[reshape_columns].rename(columns = reshape_strucs[reshape_base.value])

    if reshape_filter:

        if isinstance(reshape_filter, List):
            if not all(col in biblio_df.columns for col in reshape_filter):
                raise ValueError(f"One or more columns not found in biblio_df: {reshape_filter}")
            
            if ('bib_src' in biblio_df.columns) and ('bib_src' not in reshape_filter):
                reshape_filter.append('bib_src')

            biblio_df = biblio_df[reshape_filter]

        elif isinstance(reshape_filter, Dict):

            if not all(col in biblio_df.columns for col in reshape_filter.keys()):
                raise ValueError(f"One or more columns not found in biblio_df: {reshape_filter}")
            
            reshape_columns = list(reshape_filter.keys())

            if 'bib_src' not in reshape_columns:
                reshape_columns.append('bib_src')

            biblio_df = biblio_df[reshape_columns].rename(columns = reshape_filter)

    return biblio_df


def add_search_label(biblio_df_: pd.DataFrame,
                     search_label: str) -> pd.DataFrame:
    '''
    Add a column `search_label` to `biblio_df`.
     
    In some cases, especially when merging bibliographic datasets from different sources, it is helpful to
    know where the searches come from or what particular search terms were used. The `search_label`
    of merged datasets are merged into a list.

    Args:
        biblio_df: 
            The input bibliographic `DataFrame`.
        search_label:
            The search label that will be added to the column 'search_label.
    
    Returns:
        The bibliographic dataset `biblio_df` with the new column `search_label`.
    '''

    biblio_df = biblio_df_.copy()

    biblio_df['search_label'] = search_label
    
    return biblio_df


def add_biblio_source(biblio_df_: pd.DataFrame,
                      biblio_source: BiblioSource = BiblioSource.UNDEFINED
                      ) -> pd.DataFrame:
    """
    Adds a bibliography source column to `biblio_df` based on the specified `BiblioSource`.

    Use this function in case the `bib_src` has not been set correctly or the bibliographic
    dataset wasn't read with `read_biblio_csv_files_to_df`.

    Args:
        biblio_df: 
            The bibliographic dataset to add the bibliography source column to.
        biblio_source: 
            The bibliography source: `SCOPUS`, `LENS`, `DIMS`, `BIBLIO`. See `config.py` for 
            an updated list.

    Returns:
        pd.DataFrame: The modified DataFrame with the bibliography source column added.
    """

    biblio_df = biblio_df_.copy()

    if biblio_source == BiblioSource.SCOPUS:
        biblio_df['bib_src'] = 'scopus'
    elif biblio_source == BiblioSource.LENS:
        biblio_df['bib_src'] = 'lens'
    elif biblio_source == BiblioSource.DIMS:
        biblio_df['bib_src'] = 'dims'
    elif biblio_source == BiblioSource.BIBLIO:
        biblio_df['bib_src'] = 'biblio'

    return biblio_df


def format_auth_scopus(authors: Union[str, float]) -> Union[str, float]:
    """
    Reformats the author names string of a Scopus publication in a CSV export.

    The normalised format is

    - For a single author: `Surname, Initials` (e.g. Smith, J.A.)
    - For multipe authors: `Surname1, Initials1; Surname 2, Initials2 (e.g. Martínez, X.; Müller, H.W.)

    Example of a Scopus author names string: `Alexandre M., Silva T.C., Michalak K., Rodrigues F.A.`

    Args:
        authors: 
            The author names as a string in Scopus format. Handles NaN via `float`.

    Returns:
        The formatted author names as a string or NaN if missing.
    """

    if isinstance(authors, str) and authors != '':
        authors_lst = authors.split(',')

        authors_str_lst = []

        for author in authors_lst:

            # Split author name into parts (surname and initials)
            author_parts = author.strip().split()
            surname = author_parts[0].strip()
            initials = [name.strip()[0] + '.' if not name.endswith('.') else name.strip() for name in author_parts[1:]]
            initials = ''.join(initials)

            # Create normalised author string
            author_str = '{}, {}'.format(surname, initials)
            authors_str_lst.append(author_str)

        normalised_str = '; '.join(authors_str_lst)

        return normalised_str
    else:
        return np.nan


def format_auth_lens(authors: Union[str, float]) -> Union[str, float]:
    """
    Reformats the author names string of a Lens publication in a CSV export.

    The normalised format is

    - For a single author: `Surname, Initials` (e.g. Smith, J.A.)
    - For multipe authors: `Surname1, Initials1; Surname 2, Initials2 (e.g. Martínez, X.; Müller, H.W.)

    Example of a Lens author names string: `Elena L. Mishchenko; A. M. Mishchenko; Vladimir A. Ivanisenko`

    Args:
        authors: 
            The author names as a string in Lens format. Handles NaN via `float`.

    Returns:
        The formatted author names as a string or NaN if missing.
    """

    if isinstance(authors, str) and authors != '':
        authors_lst = authors.split(';')

        authors_str_lst = []

        for author in authors_lst:

            # Split author name into parts (surname and initials)
            author_parts = author.strip().split()
            surname = author_parts[-1]
            initials = [name[0] + '.' if not name.endswith('.') else name for name in author_parts[:-1]]
            initials = ''.join(initials)

            # Create normalised author string
            author_str = '{}, {}'.format(surname.strip(), initials.strip())
            authors_str_lst.append(author_str)

        normalised_str = '; '.join(authors_str_lst)

        return normalised_str
    else:
        return np.nan


def format_auth_dims(authors: Union[str, float]) -> Union[str, float]:
    """
    Reformats the author names string of a Dimensions publication in a CSV export.

    The normalised format is

    - For a single author: `Surname, Initials` (e.g. Smith, J.A.)
    - For multipe authors: `Surname1, Initials1; Surname 2, Initials2 (e.g. Martínez, X.; Müller, H.W.)

    Example of a Dimensions author names string: `Bai, Xiao; Sun, Huaping; Lu, Shibao; Taghizadeh-Hesary, Farhad`

    Args:
        authors: 
            The author names as a string in Dimensions format. Handles NaN via `float`.

    Returns:
        The formatted author names as a string or NaN if missing.
    """

    if isinstance(authors, str) and authors != '':
        authors_lst = authors.split(';')

        authors_str_lst = []

        for author in authors_lst:

            # Split author name into parts (surname and initials)
            author = author.replace(',', '')
            author_parts = author.strip().split()
            surname = author_parts[0].strip()
            initials = [name.strip()[0] + '.' if not name.endswith('.') else name.strip() for name in author_parts[1:]]
            initials = ''.join(initials)

            # Create normalised author string
            author_str = '{}, {}'.format(surname, initials)
            authors_str_lst.append(author_str)

        normalised_str = '; '.join(authors_str_lst)

        return normalised_str
    else:
        return np.nan


def format_auth_affils_scopus(auth_affils_str: Union[str, float]) -> Union[str, float]:
    """
    Reformats the authors-affilitions string of a Scopus publication in a CSV export.

    The normalised format is

    - For a single author-affiliation: `Surname, Initials (Affiliation(s))` (e.g. Jin, X. (Luxembourg School of Finance, University of Luxembourg, Luxembourg))
    - For multipe authors-affiliations: `Surname1, Initials1 (Affiliation(s)1); Surname 2, Initials2 (Affiliation(s)2)

    Example of a Scopus authors-affiliations names string: `Blundell-Wignall, A., Enterprise Affairs, OECD, France; Roulet, C., Enterprise Affairs, OECD, France`

    !!! note
        Since Lens does not provide affiliation information, there is no corresponding function `format_auth_affils_lens`.

    Args:
        auth_affils_str: 
            The authors-affiliations as a string in Scopus format. Handles NaN via `float`.

    Returns:
        The formatted authors-affiliations string or NaN if missing.
    """
 
    auth_affils =[]

    if isinstance(auth_affils_str, str):  # prevents a Pylance error in x.split()
        if auth_affils_str != '':
            # Split auth_affils_str into individual author-affiliation pairs
            auth_affils = [re.split(r'\.\s*,|.\s+,\s*', name.strip(), maxsplit = 1) for name in auth_affils_str.split(';')]

            # Format each author-affiliation pair
            auth_affils = [[parts[0] + '.' if len(parts) > 1 else parts[0], parts[1].strip()] if len(parts) > 1 else 
                        [parts[0].strip()] for parts in auth_affils]
        else:
            # If auth_affils_str is an empty string, return a default anonymous author-affiliation
            return 'Anonymous ()'
    else:
        return np.nan

    auth_affil_str_lst = []

    for auth_affil in auth_affils:
        
        # Create the formatted author-affiliation string in the normalised format
        # If author_affil has just one element (author or affiliation), assume that the affiliation rather than the author name is missing
        if len(auth_affil) == 1:
            auth_affil_str = '{} ()'.format(auth_affil[0].strip())
        else:
            auth_affil_str = '{} ({})'.format(auth_affil[0].strip(), auth_affil[1].strip())
        
        auth_affil_str_lst.append(auth_affil_str)

    normalised_str = '; '.join(auth_affil_str_lst)

    return normalised_str


def format_auth_affils_dims(auth_affils_str: Union[str, float]) -> Union[str, float]:
    """
    Reformats the authors-affilitions string of a Dimensions publication in a CSV export.

    The normalised format is

    - For a single author-affiliation: `Surname, Initials (Affiliation(s))` (e.g. Jin, X. (Luxembourg School of Finance, University of Luxembourg, Luxembourg))
    - For multipe authors-affiliations: `Surname1, Initials1 (Affiliation(s)1); Surname 2, Initials2 (Affiliation(s)2)

    Example of a Scopus authors-affiliations names string: `Schweizer, Pia‐Johanna (Institute for Advanced Sustainability Studies (IASS), Potsdam, Germany); Goble, Robert (Clark University, Worcester, MA, USA)`

    !!! note
        Since Lens does not provide affiliation information, there is no corresponding function `format_auth_affils_lens`.

    Args:
        auth_affils_str: 
            The authors-affiliations as a string in Dimensions format. Handles NaN via `float`.

    Returns:
        The formatted authors-affiliations string or NaN if missing.
    """

    # The auth_affil strings are split at ';' except if the ';' is inside of parentheses. In that case, it
    # separates multiple affiliations and is therefore ignored.
    if isinstance(auth_affils_str, str):  # prevents a Pylance error in x.split()
        auth_affils = [part.strip().split('(', maxsplit = 1) for part in re.split(r';(?![^()]*\))', auth_affils_str.strip())]
    else:
        return np.nan

    auth_affil_str_lst = []

    for auth_affil in auth_affils:

        # Extract the affiliation, if present
        if len(auth_affil) == 2:
            affiliation = auth_affil[1][:-1] if auth_affil[1].endswith(')') else auth_affil[1]

        else:
            affiliation = ''

        # Extract the author information, if present
        if auth_affil[0]:
            author_parts = auth_affil[0].split(',', maxsplit=1)
            surname = author_parts[0].strip()
            name_parts = [part.strip() for part in author_parts[1].split()]

            # Format the author name using the first letter of each name part
            initials = [part[0] + '.' if not part.endswith('.') else part for part in name_parts]
            name = ''.join(initials)
        else:
            # No author information
            surname = ''
            name = ''

        # Construct the normalised author-affiliation string
        auth_affil_str = np.nan

        if surname:
            auth_affil_str = f'{surname}, {name} ({affiliation})'
        elif affiliation:
            auth_affil_str = f'Anonymous ({affiliation})'

        auth_affil_str_lst.append(auth_affil_str)

    if not auth_affil_str_lst:
        return np.nan
    else:
        normalised_str = '; '.join(auth_affil_str_lst)
        return normalised_str


def get_biblio_source_string(biblio_source: BiblioSource) -> Union[str, float]:
    """
    Retrieves the string representation of a bibliography source.

    Args:
        biblio_source: The source of the bibliography (e.g. `BiblioSource.SCOPUS`).

    Returns:
        The string representation of the bibliography source (e.g. `scopus`) or NaN if not recognized.
    """

    biblio_source_str = np.nan

    if biblio_source == BiblioSource.SCOPUS:
        biblio_source_str = 'scopus'
    elif biblio_source == BiblioSource.LENS:
        biblio_source_str = 'lens'
    elif biblio_source == BiblioSource.DIMS:
        biblio_source_str = 'dims'
    elif biblio_source == BiblioSource.BIBLIO:
        biblio_source_str = 'biblio'
    else:
        raise ValueError(f"The value passed for biblio_source is invalid.")
    
    return biblio_source_str


def normalise_biblio_entities(biblio_df_: pd.DataFrame,
                              biblio_source: BiblioSource = BiblioSource.UNDEFINED
                              ) -> pd.DataFrame:
    """
    Converts certain values in the bibliographic dataset `biblio_df` to a normalised 
    format so that datasets from different bibliographic databases can be merged.

    The exact values that are normalised depend on the source (Scopus, Lens,...) of 
    the dataset provided by 'biblio_source`. If `biblio_source` is not provided, the
    function will try to obtain it from `bib_src`. This only works if there is just 
    one bibliographic source provided in `bib_src` and it is the same for are all
    publications in `biblio_df`.

    Args:
        biblio_df: 
            The bibliographic dataset.
        biblio_source: 
            The bibliographic database that is the source of the bibliographic dataset.

    Returns:
        The bibliographic dataset with normalised bibliographic entities.
    """

    biblio_df = biblio_df_.copy()

    if biblio_source == BiblioSource.UNDEFINED:

        # Check if the bib_src column exists and has consistent values
        if 'bib_src' in biblio_df.columns:

            # Determine the bibliography source based on the first record
            if (biblio_df['bib_src'].isin(['scopus', 'lens', 'dims', 'biblio'])).all():
                if biblio_df.loc[0, 'bib_src'] == 'scopus':
                    biblio_source = BiblioSource.SCOPUS
                elif biblio_df.loc[0, 'bib_src'] == 'lens':
                    biblio_source = BiblioSource.LENS
                elif biblio_df.loc[0, 'bib_src'] == 'dims':
                    biblio_source = BiblioSource.DIMS
                elif biblio_df.loc[0, 'bib_src'] == 'biblio':
                    biblio_source = BiblioSource.BIBLIO
            else:
                raise ValueError(f"The bib_src needs to be the same for all records in biblio_df")
        else:
            raise ValueError(f"You either need to pass a biblio_source to the function or add a bib_src column")

    # Create the clickable link and the list of all links
    if biblio_source == BiblioSource.SCOPUS:
        if 'doi' in biblio_df.columns:
            biblio_df['link'] = 'https://dx.doi.org/' + biblio_df['doi']
            biblio_df['links'] = biblio_df['link']
        else:
            biblio_df['link'] = np.nan
            biblio_df['links'] = np.nan
    elif biblio_source == BiblioSource.LENS:
        biblio_df['link'] = np.nan
        biblio_df['links'] = np.nan

        if 'ext_url' in biblio_df.columns:
            biblio_df['link'] = biblio_df['ext_url']
        if 'source_urls' in biblio_df.columns:
            biblio_df['links'] = biblio_df['source_urls']
            
    elif biblio_source == BiblioSource.DIMS:
        if 'ext_url' in biblio_df.columns:
            biblio_df['link'] = biblio_df['ext_url']
            biblio_df['links'] = biblio_df['ext_url']
        else:
            biblio_df['link'] = np.nan
            biblio_df['links'] = np.nan

    # Create the unified keywords lists in column 'kws'
    if biblio_source == BiblioSource.SCOPUS:
        biblio_df['kws'] = biblio_df.apply(
            lambda row: '; '.join(sorted(set(
                [kw.strip() for kw in row[['kws_author', 'kws_index']].str.cat(sep = ';', na_rep = '').lower().split(';') if kw.strip()]
            ))), axis=1)
            
    elif biblio_source == BiblioSource.LENS:
        biblio_df['kws'] = biblio_df.apply(
            lambda row: '; '.join(sorted(set(
                [kw.strip() for kw in row[['kws_lens', 'mesh']].str.cat(sep = ';', na_rep = '').lower().split(';') if kw.strip()]
            ))), axis=1)
    elif biblio_source == BiblioSource.DIMS:
        if 'mesh' in biblio_df.columns:
            # biblio_df['kws'] = biblio_df['kws'].apply(lambda x: '; '.join(sorted(x.split('; '))))
            biblio_df['kws'] = biblio_df['mesh'].apply(lambda x: '; '.join(sorted(x.lower().split('; '))) 
                                                       if pd.notna(x) and ';' in x else x.lower() if pd.notna(x) else np.nan)
        else:
            biblio_df['kws'] = np.nan

    # Normalise the authors format
    if 'authors' in biblio_df.columns:
        if biblio_source == BiblioSource.SCOPUS:
            biblio_df['authors'] = biblio_df['authors'].apply(format_auth_scopus)
        elif biblio_source == BiblioSource.LENS:
            biblio_df['authors'] = biblio_df['authors'].apply(format_auth_lens)
        elif biblio_source == BiblioSource.DIMS:
            biblio_df['authors'] = biblio_df['authors'].apply(format_auth_dims)

    # Normalise the author-affiliations format
    if 'auth_affils' in biblio_df.columns:
        if biblio_source == BiblioSource.SCOPUS:
            biblio_df['auth_affils'] = biblio_df['auth_affils'].apply(format_auth_affils_scopus)
        elif biblio_source == BiblioSource.DIMS:
            biblio_df['auth_affils'] = biblio_df['auth_affils'].apply(format_auth_affils_dims)

    # Convert some other entities to lower case
    if 'fos' in biblio_df.columns:
        biblio_df['fos'] = biblio_df['fos'].apply(lambda x: x.lower() if pd.notnull(x) else x) # biblio_df['fos'].str.lower()
    if 'kws_lens' in biblio_df.columns:
        biblio_df['kws_lens'] = biblio_df['kws_lens'].apply(lambda x: x.lower() if pd.notnull(x) else x) # biblio_df['kws_lens'].str.lower()
    if 'mesh' in biblio_df.columns:
        biblio_df['mesh'] = biblio_df['mesh'].apply(lambda x: x.lower() if pd.notnull(x) else x) # biblio_df['mesh'].fillna('').str.lower()
    if 'kws_author' in biblio_df.columns:
        biblio_df['kws_author'] = biblio_df['kws_author'].apply(lambda x: x.lower() if pd.notnull(x) else x) # biblio_df['kws_author'].fillna('').str.lower()
    if 'kws_index' in biblio_df.columns:
        biblio_df['kws_index'] = biblio_df['kws_index'].apply(lambda x: x.lower() if pd.notnull(x) else x) # biblio_df['kws_index'].fillna('').str.lower()

    return biblio_df


def remove_title_duplicates(biblio_df_: pd.DataFrame) -> pd.DataFrame:
    """
    Removes duplicate publications by title from `biblio_df`.

    The function takes a bibliographic dataset `biblio_df` as input and performs several 
    operations to remove duplicate titles. It returns the modified `DataFrame` after removing 
    duplicates.

    The major processing steps are:

    - Group all duplicate titles
    - For each group
        - Merge keywords, links, bib_src, and other columns
        - Ignore the publications with missing abstracts (unless they are all missing)
        - If there are Scopus publications, use their title, abstract, and year
        - Pick the publication with the latest publish date/year. If there are several,
          then pick one at random.

    The function merges:
    
    - bib_src (the bibliographic sources)
    - source (the publication sources, e.g. journals, conference proceedings,...)
    - n_cited (number of citations)
    - fos (fields of study; only Lens)
    - anzsrc_2020 (Australian and New Zealand Standard Research Classification)
    - kws (keywords)
    - links (URLs to access the publication)

    Args:
        biblio_df: The bibliographic dataset.

    Returns:
        The modified `DataFrame` after removing duplicates.

    Assumptions:
        The titles have previously been cleaned using the function `clean_biblio_df`. If not, there is a chance that duplicate titles are being missed because of differences in lower/upper case, special characters, etc.
    """

    biblio_df = biblio_df_.copy()

    logger.info(f"Number of publications before removing duplicate titles: {biblio_df.shape[0]}")

    # Convert the years to int (Lens has the year as a float, so I force missing values there to zero)
    biblio_df['year'] = biblio_df['year'].fillna(0).astype(int)

    # Create column pub_date if it doesn't exist (in Scopus for instance)
    if 'pub_date' not in biblio_df.columns:
        biblio_df['pub_date'] = np.nan

    # Replace the missing pub_date with 1/1/1700
    biblio_df['pub_date_dummy'] = biblio_df['pub_date'].fillna(pd.to_datetime('01-01-1700', format = '%d-%m-%Y'))

    # Remove entries where no date information is provided
    # biblio_df = biblio_df.loc[
    #     ((biblio_df['bib_src'] == 'scopus') & (biblio_df['year'] != 0)) |
    #     ((biblio_df['bib_src'] != 'scopus') & ((biblio_df['year'] != 0) | (biblio_df['pub_date_dummy'] != pd.to_datetime('01-01-1700', format = '%d-%m-%Y'))))
    # ]

    if not biblio_df.empty:

        # Create a pub_date of 01/01/year for all missing pub_dates
        biblio_df['pub_date_dummy'] = biblio_df['pub_date']
        biblio_df['pub_date_dummy'].fillna(pd.to_datetime('01-01-' + biblio_df.loc[biblio_df['year'] != 0, 'year'].astype(int).astype(str), format = '%d-%m-%Y'), inplace=True)

        # Convert all 'pub_date_dummy' values to datetime
        biblio_df['pub_date_dummy'] = pd.to_datetime(biblio_df['pub_date_dummy'], errors = 'coerce')

        # Extract missing years from pub_date
        if (biblio_df['year'] == 0).any():
            biblio_df.loc[biblio_df['year'] == 0, 'year'] = \
                biblio_df.loc[biblio_df['year'] == 0, 'pub_date_dummy'].dt.year

    # Create new columns for merged values
    biblio_df['bib_srcs'] = biblio_df['bib_src']

    if 'source' in biblio_df:
        biblio_df['sources'] = np.nan
    
    # Dataframe for duplicate titles
    dup_df = pd.DataFrame()

    # Group the rows by the titles and loop over the groups
    for idx, (_, group) in enumerate(biblio_df.groupby('title')):

        # Create a new column 'sources' that has all the source titles
        if 'source' in group.columns:
            unique_src = group['source'].dropna().str.split(';').explode().str.strip().unique()
            group['sources'] = '; '.join(unique_src)
            biblio_df.loc[group.index, 'sources'] = group['source']

        if(idx % 1 == 0):
            print(f'Duplicate group: #{idx} ', end = '\r')

        if len(group) > 1:
            dup_df = pd.concat([dup_df, group])

            # Identify the duplicates by index within this group
            duplicates = group.duplicated(subset = 'title', keep = False)

            # Sum the values in n_cited
            if 'n_cited' in group.columns:
                sum_n_cited = group['n_cited'].sum()
                group['n_cited'] = sum_n_cited
                biblio_df.update(group[['n_cited']])
            
            # Merge the values in the fos column
            if 'fos' in group.columns:
                unique_fos = group['fos'].dropna().str.split(';').explode().str.strip().unique()
                group['fos'] = '; '.join(unique_fos)
                biblio_df.update(group[['fos']])

            # Merge the values in the anzsrc_2020 column
            if 'anzsrc_2020' in group.columns:
                unique_anzsrc = group['anzsrc_2020'].dropna().str.split(';').explode().str.strip().unique()
                group['anzsrc_2020'] = '; '.join(unique_anzsrc)
                biblio_df.update(group[['anzsrc_2020']])

            # Merge the values in the keywords (kws) column
            if 'kws' in group.columns:
                unique_kws = group['kws'].str.lower().dropna().str.split(';').explode().str.strip().unique()
                group['kws'] = '; '.join(unique_kws)
                biblio_df.update(group[['kws']])

            # Create a new column 'links' that has all the URLs separate with a space
            if 'links' in group.columns:
                unique_links = group['links'].dropna().str.split(' ').explode().str.strip().unique()
                group['links'] = ' '.join(unique_links)
                biblio_df.update(group[['links']])

            # Create a new column 'bib_srcs' that has all the bib_src strings of the duplicate titles
            unique_bib_src = group['bib_src'].dropna().str.split(',').explode().str.strip().unique()
            group['bib_srcs'] = ', '.join(unique_bib_src)
            biblio_df.update(group[['bib_srcs']])

            # Pick an author affiliation string
            if 'author_affils' in group.columns:
                has_scopus = group['bib_src'].str.contains('scopus', case = False, na = False)
                filtered_group = group['author_affils'].dropna()

                if has_scopus.any():
                    filtered_group = filtered_group.loc[has_scopus]
                
                if not filtered_group.empty:
                    selected_row = filtered_group.sample(n = 1)
                    group['author_affils'] = selected_row.iloc[0]

            # If at least one publication in the group has an abstract,
            # remove any publication in the group that doesn't have an abstract
            # and then reset the group with the removed publications
            if group['abstract'].notna().any():
                items_to_drop = group[duplicates & group['abstract'].isna()]    # drops all items except the item_to_keep
                group = group.drop(items_to_drop.index)
                biblio_df = biblio_df.drop(items_to_drop.index)

            # Check if any publication in the group has bib_src = 'scopus'
            if 'scopus' in group['bib_src'].values:

                # Identify the duplicates within this group
                duplicates = group.duplicated(subset = 'title', keep = False)
                scopus_mask = (group['bib_src'] == 'scopus')

                scopus = group[scopus_mask]

                if scopus.shape[0] > 0:
                    latest_scopus = scopus.loc[[scopus['year'].idxmax()]]

                    if latest_scopus['abstract'].isna().any():
                        scopus_with_abstract = scopus.dropna(subset = ['abstract'], how = 'any')

                        if scopus_with_abstract.shape[0] > 0:
                            item_to_keep = scopus_with_abstract.sample()
                        else:
                            item_to_keep = scopus.sample()
                    else:
                        item_to_keep = latest_scopus

                    biblio_df = biblio_df.loc[biblio_df['bib_src'] == 'scopus']
                    group = group.loc[group['bib_src'] == 'scopus']

                    items_to_drop = scopus.drop(item_to_keep.index) # drops all items except the item_to_keep
                    group = group.drop(items_to_drop.index)
                    biblio_df = biblio_df.drop(items_to_drop.index)

            # If there are several publications left, select the one with the latest
            # pub_date. If they are all the same, select one randomly.
            if len(group) > 1:

                # If there are multiple records with the same largest pub_date, first check whether 
                # there is one that has a source. If there are multiple, then pick one at random
                if 'source' in group.columns and group['source'].notna().any():
                    idxmax_group = group[group['source'].notna()]
                else:
                    idxmax_group = group.loc[group['pub_date_dummy'] == group['pub_date_dummy'].max()]

                idxmax_values = idxmax_group.index.values
                random_idx = np.random.choice(idxmax_values)

                # Drop all rows except for the one with the largest pub_date
                indices_to_drop = group.index.difference([random_idx])
                group = group.drop(indices_to_drop)
                biblio_df = biblio_df.drop(indices_to_drop)

        else:
            # Copy single values to new columns
            biblio_df['bib_srcs'] = biblio_df['bib_src']

    # Remove the dummy column
    biblio_df = biblio_df.drop(['pub_date_dummy'], axis = 1)

    # Change n_cited to int
    if 'n_cited' in biblio_df.columns:
        biblio_df['n_cited'] = biblio_df['n_cited'].fillna(0)
        biblio_df['n_cited'] = biblio_df['n_cited'].astype(int)

    logger.info(f"Number of publications after removing duplicate titles: {biblio_df.shape[0]}")

    return biblio_df


def clean_biblio_df(biblio_df_: pd.DataFrame) -> pd.DataFrame:
    """
    Cleans the bibliographic dataset `biblio_df` by processing the publication titles
    and abstracts to remove noise and standardise formatting.

    The function removes publications with empty titles, NaN titles, and titles containing 
    specific keywords such as 'conference', 'workshop', 'proceedings'. It converts titles to 
    lowercase, removes HTML tags, non-alphabetic characters (except specific punctuation and 
    Greek letters), excess whitespace, and common terms from the beginning of titles.

    For abstracts, the function replaces NaN values with empty strings, removes HTML tags, 
    non-alphabetic characters (except specific punctuation and Greek letters), excess whitespace, 
    and common terms used in abstracts of publications in some disciplines such as 'background', 
    'objectives', 'results' etc. It also removes duplicate publications using the function 
    `remove_title_duplicates` and generates standard publication IDs.

    Args:
        biblio_df: 
            Input DataFrame containing bibliographic information.

    Returns:
        Cleaned bibligraphic dataset with standardized publication titles, abstracts, and new publication IDs.

    Raises:
        None.

    """

    biblio_df = biblio_df_.copy()

    logger.info(f'Number of publications in the input biblio_df: {len(biblio_df)}')


    """
        Cleaning the publication titles
    """

    # Remove publications with empty titles
    count_titles_empty = len(biblio_df[biblio_df['title'] == ''])
    biblio_df = biblio_df[biblio_df['title'] != '']
    print(f'Removed {count_titles_empty} titles that were empty strings')

    # Remove all titles that are NaN
    count_titles_nan = biblio_df['title'].isna().sum()
    biblio_df = biblio_df.dropna(subset = ['title'])
    print(f'Removed {count_titles_nan} titles that were NaN')

    # Remove all titles that contain 'conference', 'workshop', or 'proceedings'
    count_procs = len(biblio_df[biblio_df['title'].str.contains('proceedings|conference|workshop', case = False)])
    biblio_df = biblio_df[~biblio_df['title'].str.contains('proceedings|conference|workshop', case = False)]
    print(f'Removed {count_procs} records where the title contained "conference", "workshop", or "proceeding"')

    # Convert the titles to lower case except for the first word
    def title_to_lc(s: str) -> str:
        clean_s = re.sub(r'[^A-Za-z]', ' ', s)
        if clean_s.isupper():
            final_s = clean_s.capitalize()
        else:
            words = s.split()
            words = [words[0]] + [w.lower() if w[0].isupper() and len(w) > 1 and w[1].isalpha() and not w[1].isupper() else w for w in words[1:]]
            final_s = ' '.join(words)
        return final_s

    biblio_df['title'] = biblio_df['title'].apply(title_to_lc)

    # Remove text between '<' and '>' characters
    biblio_df['title'] = biblio_df['title'].str.replace(r'<.*?>', '', regex = True)

    # Remove all non-alphabetic characters except for '-', ',', ':', ')', '(', '$', '%',  whitespace and Greek letters
    biblio_df['title'] = biblio_df['title'].apply(lambda x: re.sub(r'[^a-zA-Z0-9α-ωΑ-Ω\s,:’()$%\'\"\-]+', ' ', x))

    # Replace any special whitespace character with a normal whitespace
    biblio_df['title'] = biblio_df['title'].str.replace(r'\u2002|\u2003|\u2005|\u2009|\u200a|\u202f|\xa0', ' ', regex = True)

    # Replace all newline and tab characters with a white space
    biblio_df['title'] = biblio_df['title'].str.replace(r'\n|\t', ' ', regex = True)

    # Remove the word 'abstract' at the start of any title
    biblio_df['title'] = biblio_df['title'].str.replace(r'(?i)^abstract\s*', '', regex = True)

    # Remove words from the beginning of the title that are combinations of at least one number and zero or more special charcters
    biblio_df['title'] = biblio_df['title'].apply(lambda x: re.sub(r'^[\W\d]+(?=\s)', '', x))

    # Remove any words starting with '.' or '-' and any single letter except 'a' from the beginning of the title and abstract
    biblio_df['title'] = biblio_df['title'].str.replace(r'^[-.]+\s*\w+\s*|[-.]+(?!\w)|(\s|^)[^Aa\s+](\s+|$)', '', regex = True)

    # Remove excess whitespace
    biblio_df['title'] = biblio_df['title'].str.replace(r'\s+', ' ', regex = True).str.strip()

    # Remove any remaining empty titles
    count_titles = biblio_df.shape[0]
    biblio_df = biblio_df[biblio_df['title'].str.strip().astype(bool)]
    print(f'Removed additional {count_titles - biblio_df.shape[0]} titles that were empty strings')


    """
        Cleaning the publication abstracts
    """

    # Replace all abstracts that are NaN with empty strings
    count_abs_nan = biblio_df['abstract'].isna().sum()
    biblio_df['abstract'] = biblio_df['abstract'].fillna('')
    print(f'Replaced {count_abs_nan} abtracts that were NaN with an empty string')

    # Remove text between '<' and '>' characters
    biblio_df['abstract'] = biblio_df['abstract'].str.replace(r'<.*?>', '', regex = True)

    # Remove all non-alphabetic characters except for '.', '-', ',', ':', ')', '(', '$', '%',  whitespace and Greek letters
    biblio_df['abstract'] = biblio_df['abstract'].apply(lambda x: re.sub(r'[^a-zA-Z0-9α-ωΑ-Ω\s.,:’()$%\'\"\-]+', ' ', x))

    # Replace any special whitespace character with a normal whitespace
    biblio_df['abstract'] = biblio_df['abstract'].str.replace(r'\u2002|\u2003|\u2005|\u2009|\u200a|\u202f|\xa0', ' ', regex = True)

    # Replace all newline and tab characters with a white space
    biblio_df['abstract'] = biblio_df['abstract'].str.replace(r'\n|\t', ' ', regex = True)

    # Remove the word 'abstract' and 'objective' at the start of any abstract
    biblio_df['abstract'] = biblio_df['abstract'].str.replace(r'(?i)^abstract\s*', '', regex = True)
    biblio_df['abstract'] = biblio_df['abstract'].str.replace(r'(?i)^objective(s)?\s*', '', regex = True)

    # Remove the following common terms from the abstract independently of the case
    remove_strings = ['background', 'objective', 'results', 'conclusions', 'introduction']
    pattern = "|".join(remove_strings)
    biblio_df['abstract'] = biblio_df['abstract'].apply(lambda x: re.sub(pattern, '', x, flags = re.IGNORECASE))

    # Remove any words starting with '.' or '-' and any single letter except 'a' from the beginning of the title and abstract
    biblio_df['abstract'] = biblio_df['abstract'].str.replace(r'^[-.]+\s*\w+\s*|(\s|^)[^Aa\s+](\s+|$)', '', regex = True)

    # Remove excess whitespace
    biblio_df['abstract'] = biblio_df['abstract'].str.replace(r'\s+', ' ', regex = True).str.strip()


    """
        Removing duplicate publications
    """

    biblio_df = remove_title_duplicates(biblio_df)
    

    # Convert year to integer and replace nan with zeros (in Lens, year is a float)
    if 'year' in biblio_df.columns and biblio_df['year'].dtype == float:
        biblio_df['year'] = pd.to_numeric(biblio_df['year'], errors = 'coerce')
        biblio_df['year'] = biblio_df['year'].fillna(0)
        biblio_df['year'] = biblio_df['year'].astype(int)


    '''
        Generate the new publication IDs
    '''
    
    # Sort the dataset before creating the ids
    biblio_df = biblio_df.sort_values(by = ['year', 'title'], ascending = [False, True], na_position='last')

    # Generate the record IDs
    global counter
    counter = 0

    def generate_id(row):
        global counter
        author = ''

        if (row['authors'] != "") and isinstance(row['authors'], str):
            if "no author name" in row['authors'].lower():
                author = 'Anonymous'
            else:
                author = row['authors'].split()[0].strip().strip(',')
        else:
            author = 'Anonymous'

        id = str(counter).zfill(6) + '_' + author + '_' + str(row['year'])
        counter += 1

        return id
    
    # Make sure the index is properly set (starts at 0, then at unit increments)
    biblio_df.reset_index(drop = True, inplace = True)

    # Generate the unique identifiers for the publications
    biblio_df['id'] = biblio_df.apply(generate_id, axis = 1)
    
    return biblio_df


def generate_pandas_query_string(query_str: str) -> str:
    """
    Generates a query string that can be used by the pandas `query` function  based on 
    the input `query_str`.

    Pandas query strings can be cumbersome, for instance when string functions need to 
    be used on columns for partial/substring matches. You then need to use strings like
    `~(title.str.contains(r'(?=.*systemic risk)^(?=.*\\bequity\\b)', case = False, regex = True))`,
    which are unwieldy to create.

    This function allows queries to be written in a simplified form that covers most needs
    for column filtering in bibliographic `DataFrames`. Here are some query examples:
    
    - filter_query = "systemic in* title"
    - filter_query = "public in* title | extreme in* abstract"
    - filter_query = "013 in* lens_id"
    - filter_query = "'exists' in* abstract"
    - filter_query = "['modules', extreme] in* title"
    - filter_query = "extreme in* [title, abstract]"
    - filter_query = "[[systemic, 'equity']] in* title"
    - filter_query = "[[procyc, 'module']] in* [title, abstract]"
    - filter_query = "['modules', extreme] in* [title, abstract]"
    - filter_query = "~systemic in* title"
    - filter_query = "~(~public in* title | extreme in* abstract)"
    - filter_query = "~'exists' in* abstract"
    - filter_query = "~['modules', extreme] in* title"
    - filter_query = "~[[systemic risk, 'equity']] in* title"
    - filter_query = "(risk in* title) & (year == 2014)"

    The syntax rules are:

    - In addition to the pandas query() syntax, there is an additional binary operator `in*`.
    - The general syntax is `STRINGS in* COLUMNS`. This searches for (sub)strings in the
      text of the specified columns.
    - STRINGS:
        - If you put single or double quotes around a string, it will search for a full 
        word match. Otherwise it will search for partial matches.
        - If you surround multiple strings with single square brackets, it will search 
          for *any* matches ('or' operation). If you use double square brackets, it will
          search for *all* matches ('and' operation).
        - You can use the  ('not') operator inside the square brackets.
    - COLUMNS:
        - You can use single brackets to search in multiple columns for *any* matches ('or' 
          operation). So, `risk in* [title, abstract]` searches for the substring "risk" 
          in columns 'title' and 'abstract' and produces a match when it finds "risk" in 
          the value title or in the abstract.
        - Double brackets are not allowed for columns. You can achieve this by using `&``
          outside the `in*` operation.
        - The ~ ('not') operator is not allowed for columns.

    - Outside the `in*` binary operations, you can use the query() operators such as &, |, ~~
      in the normal way. See the last example above.

    Args:
        query_str: Input query string to be modified.

    Returns:
        Modified query string that the Pandas `query` function understands.

    Raises:
        ValueError: If the not operator '~' is used inside square brackets '[...]' or applied to the columns using 'in*'.

    """

    # Define the characters to split query string
    split_at = r"()|&~"
    pad = r"|&~"
    split_at_escaped = re.escape(split_at)

    if re.search(r'\[[^\]]*~[^\]]*\]', query_str):
        raise ValueError(f"The not operator '~' is not allowed inside square brackets '[...]'")
    
    if re.search(r'in\*\s*~', query_str):
        raise ValueError(f"Applying the not operator '~' on the columns is not allowed")
    
    # Split the query string into parts at the split_at characters
    pattern = fr"(?=[{split_at_escaped}])|(?<=[{split_at_escaped}])"
    query_parts = re.split(pattern, query_str)
    query_parts = [item.strip() for item in query_parts if item.strip()]

    modified_query_parts = []

    for query_part in query_parts:

        if 'in*' in query_part:

            # Split the 'in*' part into search string and column string
            in_parts = query_part.split('in*')
            search_str = in_parts[0].strip()
            col_str = in_parts[1].strip()                

            # Check if multiple columns are specified for searching
            match_col = re.match(r"^\[.*\]$", col_str)

            if match_col:
                matched_col_lst = [string.strip() for string in match_col.group().replace('[', '').replace(']', '').split(',')]
            else:
                matched_col_lst = [col_str]

            # Check for multiple search strings and the operation type ('and' or 'or')
            match_and = re.match(r"^\[\[.*\]\]$", search_str)
            match_or = re.match(r"^\[.*\]$", search_str)
            match = match_and if match_and else (match_or if match_or else False)

            if match:
                # Mutiple search terms
                matched_string = match.group().strip().replace('[', '').replace(']', '')
                search_str_lst = [string.strip() for string in matched_string.split(',')]
                first_col_str = True
                subst_query = ""

                for col_str in matched_col_lst: # one or several columns

                    # Build the query string for multiple search terms and one or multiple columns
                    subst_query += f"{' | ' if not first_col_str else ''}({col_str}.str.contains(r'"
                    first_search_str = True

                    for search_str in search_str_lst:
                        if re.match(r"^['\"].*['\"]$", search_str):
                            # Single search string with full word match enforced
                            search_str = search_str.strip().replace("'", "").replace('"', "")
                            if match_and:
                                subst_query += f"{'^' if not first_search_str else ''}(?=.*\\b{search_str}\\b)"
                            else:
                                subst_query += f"{'|' if not first_search_str else ''}\\b{search_str}\\b"
                            first_search_str = False
                        else:
                            # Single search string with partial word match allowed
                            search_str = search_str.strip()
                            if match_and:
                                subst_query += f"{'^' if not first_search_str else ''}(?=.*{search_str})"
                            else:
                                subst_query += f"{'|' if not first_search_str else ''}{search_str}"
                            first_search_str = False

                    subst_query += f"', case = False, regex = True))"
                    first_col_str = False

                modified_query_parts.append(subst_query)
            else:
                # Single search term
                first_col_str = True
                subst_query = ""

                for col_str in matched_col_lst: # one or several columns
 
                    # Build the query string for multiple search terms and one or multiple columns
                    subst_query += f"{' | ' if not first_col_str else '('}"

                    if re.match(r"^['\"].*['\"]$", search_str):
                        # Single search string with full word match enforced
                        search_str = search_str.strip().replace("'", "").replace('"', "")
                        subst_query += f"{col_str}.str.contains(r'\\b{search_str}\\b', case = False, regex = True)"
                    else:
                        # Single search string with partial word match allowed
                        subst_query += f"{col_str}.str.contains('{search_str}', case = False, regex = True)"
                    
                    first_col_str = False

                subst_query += f")"
                modified_query_parts.append(subst_query)
        else:
            modified_query_parts.append(query_part)

    # Add padding spaces around operators
    modified_query_parts = [s if s not in pad else f" {s} " for s in modified_query_parts]

    # Generate the query string from the individual parts
    modified_query = ''.join(modified_query_parts)

    return modified_query


def filter_biblio_df(biblio_df: pd.DataFrame, 
                     query_str: str
                     ) -> pd.DataFrame:
    """
    Filter biblio_df using a query string.

    Parameters:

    biblio_df (pd.DataFrame): 
        The DataFrame to filter.
    query_str (str): 
        The query string to use for filtering.

    Raises:
        ValueError: If the query string is not valid, or if any of the column names
        referred to in the query string do not exist in the DataFrame.

    Returns:

    pd.DataFrame: 
        The filtered DataFrame.
    """

    # Build pandas query string
    filter_pandas_query = generate_pandas_query_string(query_str = query_str)

    # Check that the query string is valid
    try:
        filtered_df = biblio_df.query(filter_pandas_query)
    except (SyntaxError, ValueError) as e:
        raise ValueError("Invalid query string") from e

    # Return the filtered DataFrame
    return filtered_df


def highlight_selected_text(text: str, 
                            strings: Union[List[str], str], 
                            colour: str, 
                            partial: bool = True, 
                            i_row: Optional[int] = None
                            ) -> str:
    text = str(text)

    if isinstance(strings, str):
        strings = [strings]
    
    if i_row and (i_row % 100 == 0) and (logger.get_level() == logging.INFO): # print row index
        print(f'{i_row}', end = '\r')

    if len(strings) == 0:
        return text
    
    if (len(strings) == 1) and (strings[0] == ""):
        return text

    for k in strings:
        if partial: # highlight partial and full matches
            # pattern = r"\b\w*{}+\w*\b".format(k)
            pattern = r"(?<!>)\b\w*{}+\w*\b(?!<)".format(k)
            text = re.sub(pattern, lambda match: f'<span style="color: {colour}; font-weight: bold">{match.group()}</span>', text, flags = re.IGNORECASE)
        else:   # highlight full matches only
            # pattern = r"(?i)\b"+k+r"[\w-]*"
            pattern = r"(?i)(?<!>)\b"+k+r"[\w-]*\b(?!<)"
            text = re.sub(pattern, lambda match: f'<span style="color: {colour}; font-weight: bold">{match.group()}</span>', text)

    return text


def colour_name_to_hex(colour_name):
    # See the colour names here: https://www.w3.org/TR/SVG11/types.html#ColorKeywords

    try:
        # Get RGB value for the color name
        rgb = webcolors.name_to_rgb(colour_name)

        # Convert RGB to hex code
        hex_code = webcolors.rgb_to_hex(rgb)

        return hex_code
    except:
        raise ValueError(f"The colour {colour_name} does not exist")


def get_excel_column(number):
    column = ""
    while number > 0:
        number, remainder = divmod(number - 1, 26)
        column = chr(65 + remainder) + column
    return column


def excel_highlights_builder(biblio_highlights_df: pd.DataFrame,
                             excel_params: Optional[Any]) -> Workbook:

    # FIXME: This needs the lxml package installed for Workbook.save to work. Add that to the requirements.txt

    excel_sheet_name = 'Highlights' # default name for the new Excel sheet with the highlights
    excel_zoom = 100    # default zoom of the sheet
    excel_freeze_panes = None   # default freeze for horizontal panes
    excel_cols = None
    excel_highlighted_cols = None

    default_width = 15
    default_wrap = False

    # Unpack the parameters
    if isinstance(excel_params, Dict):
        excel_cols = excel_params.get('cols', None)
        excel_sheet_name = excel_params.get('sheet_name', None)
        excel_zoom = excel_params.get('zoom', excel_zoom)
        excel_freeze_panes = excel_params.get('freeze_panes', excel_freeze_panes)
        excel_highlighted_cols = excel_params.get('highlighted_cols', None)
    elif excel_params != None:
        raise ValueError(f"The function argument excel_params has to be a dictionary")
    
    formatted_cols = []
    remaining_cols = []

    # Get the columns to be added first (the remaining columns are added as-is to the Excel sheet)
    if isinstance(excel_cols, List):
        formatted_cols = [d['col'] for d in excel_cols if d['col'] in biblio_highlights_df.columns]
        remaining_cols = [col for col in biblio_highlights_df.columns if col not in formatted_cols]
    elif excel_cols == None:  # if no columns have been provided in excel_params, use all columns from biblio_highlights_df
        remaining_cols = biblio_highlights_df.columns
    else:
        raise ValueError(f"The item 'excel_cols' in the dictionary has to be a list")

    # Reorder the columns in biblio_highlights_df
    reorder_cols = formatted_cols + remaining_cols
    biblio_highlights_df = biblio_highlights_df[reorder_cols]

    # Create a new workbook
    wb = Workbook()

    # Create a new sheet
    ws = wb.create_sheet(excel_sheet_name)

    # Remove the default sheet
    if wb["Sheet"]:
        wb.remove(wb["Sheet"])

    # Make a copy of titles_highlights_df
    tak_excel_df = biblio_highlights_df.copy()

    # Excel column headers
    for j, col in enumerate(formatted_cols):
        # header = str(tak_excel_df.columns[j])
        header = col

        if excel_cols is not None:
            header = next((col_info.get('heading', col) for col_info in excel_cols if col_info['col'] == col), col)

        ws.cell(row = 1, column = j + 1, value = header).font = Font(bold = True)

    cursor = len(formatted_cols)

    for j, col in enumerate(remaining_cols):
        header = col

        if excel_cols is not None:
            header = next((col_info.get('heading', col) for col_info in excel_cols if col_info['col'] == col), col)

        ws.cell(row = 1, column = j + cursor + 1, value = header).font = Font(bold = True)

    # Apply findall() to split a string at '<span.../span>'
    def split_string_at_span(string):
        lst = re.findall(r"(.*?)(<span.*?/span>|$)", string)
        lst = [elem for tup in lst for elem in tup]
        lst = [x for x in lst if x.strip()]
        return lst

    def replace_span_with_textblock(lst: List[str]) -> List[str]:
        is_prev_kw = False  # need to add a space between two consecutive keywords
        rich_text_lst = []

        for string in lst:
            if string.startswith('<span'):
                colour_name = re.findall(r'color: (\w+)', string)[0] if re.findall(r'color: (\w+)', string) else None
                colour_hex = '00' + colour_name_to_hex(colour_name = colour_name)[1:]
                # span_string = '<span style="color: {}; font-weight: bold">'.format(colour_name)
                text = (' ' if is_prev_kw else '') + string.split('>')[1].split('<')[0]
                text_block = TextBlock(InlineFont(b = True, color = colour_hex), text)  # type: ignore (the Pylance issue is caused by the TextBlock constructor typing)
                rich_text_lst.append(text_block)
                is_prev_kw = True
            else:
                rich_text_lst.append(string)
                is_prev_kw = False

        return rich_text_lst

    if excel_highlighted_cols is not None:
        excel_highlighted_cols = [col for col in excel_highlighted_cols if col in biblio_highlights_df.columns]

        for col in excel_highlighted_cols:
            # Create a list of each cell content by splitting at '<span...>some text</span>' using findall()
            tak_excel_df[col] = tak_excel_df[col].apply(split_string_at_span)

            # Replace all '<span...>some text</span>' with the results of calling TextBlock(bold_red, 'some text')
            tak_excel_df[col] = tak_excel_df[col].apply(replace_span_with_textblock)

    numeric_cols = biblio_highlights_df.select_dtypes(include = "number").columns.tolist()

    # Loop through rows and columns of the dataframe
    for i in range(len(tak_excel_df)):
        for col in tak_excel_df.columns:
            j = tak_excel_df.columns.get_loc(col)

            if col in excel_highlighted_cols:
                rs = CellRichText(tak_excel_df.iloc[i, j])  # type: ignore - not sure how to avoid the Pylance typing warning
            elif col in numeric_cols:
                # rs = str(tak_excel_df.iloc[i, j])
                rs = tak_excel_df.iloc[i, j]
            else:
                rs = str(tak_excel_df.iloc[i, j])

            ws.cell(row = i + 2, column = j + 1, value = rs)

    for idx, col in enumerate(formatted_cols):
        col_letter = get_excel_column(idx + 1)
        width = default_width
        wrap = default_wrap

        if excel_cols is not None:
            width = next((col_info.get('width', default_width) for col_info in excel_cols if col_info['col'] == col), default_width)
            wrap = next((col_info.get('wrap', default_wrap) for col_info in excel_cols if col_info['col'] == col), default_wrap)

        ws.column_dimensions[col_letter].width = width

        for cell in ws[col_letter]:
            cell.alignment = Alignment(wrap_text = wrap)

    ws.sheet_view.zoomScale = excel_zoom

    if excel_freeze_panes:
        ws.freeze_panes = excel_freeze_panes

    return wb


def highlight_keywords(biblio_df: pd.DataFrame,
                        highlight_params: List[Dict[str, Union[List[str], str]]],
                        html_cols: Optional[List] = None,
                        xlsx_cols: Optional[List] = None,
                        excel_params: Optional[Any] = None
                        ) -> Tuple[Optional[HTML], Optional[Workbook]]:
    biblio_highlights_df = biblio_df.copy()

    highlighted_cols = []

    for param in highlight_params:
        if not all([key in param for key in ['strings', 'targets']]):
            raise ValueError(f"The key 'strings' or 'targets' is missing in the parameter dictionary")
        
        if 'colour' not in param:
            colour = 'red'
        else:
            colour = param.get('colour')

        if isinstance(param['strings'], list):
            strings = param.get('strings', [])
        else:
            strings = str(param.get('strings'))

        if isinstance(param['targets'], list):
            targets = param.get('targets', [])
        else:
            targets = [param.get('targets')]

        # Highlight the strings provided in param
        for target in targets:
            if target in biblio_highlights_df.columns:
                highlighted_cols.append(target)
                print(f'Highlighting strings in {colour} in the {target}...')
                biblio_highlights_df[target] = biblio_highlights_df \
                    .apply(lambda x: highlight_selected_text(
                            text = x[target], 
                            strings = strings, 
                            colour = str(colour), 
                            i_row = biblio_highlights_df.index.get_loc(x.name)), axis = 1)

    html_out = None
    xlsx_out = None

    if html_cols != None:
        if html_cols == []:
            html_out = HTML(biblio_highlights_df.to_html(escape = False))
        else:
             html_out = HTML(biblio_highlights_df[html_cols].to_html(escape = False))

    if excel_params != None:
        excel_params['highlighted_cols'] = list(set(highlighted_cols))

    if xlsx_cols != None:
        if xlsx_cols == []:
            xlsx_cols = list(biblio_highlights_df.columns)
        
        xlsx_out = excel_highlights_builder(biblio_highlights_df = biblio_highlights_df[xlsx_cols],
                                            excel_params = excel_params)
    else:
        xlsx_out = excel_highlights_builder(biblio_highlights_df = biblio_highlights_df,
                                            excel_params = excel_params)

    return html_out, xlsx_out


